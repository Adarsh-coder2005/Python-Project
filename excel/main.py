from openpyxl import load_workbook
workbook = load_workbook(filename='Salary_Data.xlsx')
workbook.sheetnames

sheet = workbook.active
print(sheet)

print(sheet.title)

print(sheet['A1'])

print(sheet['A1'].value)

print(sheet['B1'].value)

print(sheet.cell(row=12,column=1))

print(sheet.cell(row=12,column=1).value)

print(sheet['A:B'])

print(sheet[5])


